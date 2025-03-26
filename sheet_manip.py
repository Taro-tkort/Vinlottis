from openpyxl import load_workbook
import numpy as np
import random

print("Starting sheet manipulator\n")


def choose_random_values(nr_of_people, x):
    # Ensure x is not greater than nr_of_people
    x = min(int(x), int(nr_of_people))
    
    # Choose x random values between 0 and nr_of_people
    random_values = random.sample(range(nr_of_people), x)
    
    return random_values

## FORMULAS ##

#calculate ticket yeild-rate
def yeild(tickets):
    rate = ratio/sum
    stickets = tickets*rate
    print("Calculating rate with: Tickets: %d    Sum: %d    Rate: %d     Total: %d " % (tickets, sum, rate, stickets))
    return stickets



## Code ##

path = "Lodd.xlsx"
workbook = load_workbook(filename=path)
sheet = workbook.active

#purge the xlsx of all mt rows >:}
rowstop = 0
for row in sheet.iter_rows():
    print(row[0].value)
    if row[0].value is None or row[0].value == "":
        rowstop = row[0].row
        break
print(rowstop)
print(sheet.max_row)
sheet.delete_rows(rowstop, sheet.max_row)
workbook.save("Lodd.xlsx")
print("Purged the empty rows")

sheet['D4'] = "test"

data = [[],[],[],[],[]] #list for the extracted data
#extract relevat data from sheet
for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
    data[0].append(row[0]) #name
    data[1].append(row[1]) #money 
    print(row)

#convert money to tickets
for val in data[1]:
    tmp = val/20 #convert to tickets
    data[2].append(tmp) #add to data 3 thingy

#calculate total points
ratio = 200
sum = 0
for val in data[2]:
    sum += val

new_tickets_per_person,random_tickets_to_distribute=divmod(200-sum,len(data[0]))


for val in data[2]:
    tmp = yeild(val)
    data[3].append(val+int(new_tickets_per_person))
    
for val in data[3]:
    data[4].append(val)



nr_of_people=len(data[0])

extra_tickets_to_add=choose_random_values(nr_of_people,random_tickets_to_distribute)

print(extra_tickets_to_add)


for val in extra_tickets_to_add: 
     data[4][val]=data[4][val]+1

print(data[2])
print(data[3])
print(data[4])


#write the tickets to the excel doc
iterator = 1 #iterator running through the for loop
rowstr = "C" #the row the range is to be set at 
for val in data[2]:
    cordtmp = rowstr + str(iterator)
    print("Writing %d to %s" % (val, cordtmp))
    sheet[cordtmp]=val
    iterator += 1

#write additional tickets for each person to the excel doc
iterator = 1 #iterator running through the for loop
rowstr = "D" #the row the range is to be set at 
for val in data[3]:
    cordtmp = rowstr + str(iterator)
    print("Writing %d to %s" % (val, cordtmp))
    sheet[cordtmp]=val
    iterator += 1

#write the randomely ditributed tickets to the excel doc
iterator = 1 #iterator running through the for loop
rowstr = "E" #the row the range is to be set at 
for val in data[4]:
    cordtmp = rowstr + str(iterator)
    print("Writing %d to %s" % (val, cordtmp))
    sheet[cordtmp]=val
    iterator += 1


#writing the different tickets (1-200)
rowstr="A"
offset=nr_of_people+1
nr_to_choose=[]
for i in range(1,201):
    cordtmp=rowstr+str(i+offset)
    sheet[cordtmp]=i
    nr_to_choose.append(i)


#writing a name for each ticket
rowstr="B"
for row in range(len(data[0])):
    for nr_of_tickets in range(int(data[4][row])):
        ticket_indeks=random.randint(0,len(nr_to_choose)-1)
        ticket=nr_to_choose[ticket_indeks]
        cordtmp=rowstr+str(ticket+offset)
        sheet[cordtmp]=data[0][row]
        nr_to_choose.pop(ticket_indeks)




print("Saving as new excel sheet")
workbook.save("Fordelte-lodd.xlsx")
print("Finished! sould be written to excel now :)")

